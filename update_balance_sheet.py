#!/usr/bin/env python3
"""
QuickBooks Journal Entry Automation

- Fetches journal entries from QuickBooks Online
- Aggregates totals per account and month (current year only)
- Updates Excel workbook accordingly

Author: Damaris Chege
"""

import os
import re
import json
import requests
from datetime import datetime
from openpyxl import load_workbook

# =========================
# CONFIG
# =========================
CONFIG_PATH = os.getenv("CONFIG_PATH", "config.json")
EXCEL_PATH = os.getenv("EXCEL_PATH", "BalanceSheetItems.xlsx")
LOG_PATH = os.getenv("LOG_PATH", "logs/automation.log")

TOKEN_URL = "https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer"

ACCOUNTS = [
    ("Debtors Outside Macheo", "2000 · Debtors Outside Macheo"),
    ("Debtors Employees Macheo", "2001 · Debtors Employees Macheo"),
    ("Cash in hands Employees", "2002 · Cash in hands Employees"),
    ("Prepaid Expenses", "2005 · Prepaid Expenses"),
    ("Creditors Employees Macheo", "2007 · Creditors Employees Macheo"),
    ("Creditors Outside Macheo", "2008 · Creditors Outside Macheo"),
    ("Statutory Payables", "2009 · Statutory Payables"),
]

CURRENT_YEAR = datetime.now().year

# =========================
# LOGGING
# =========================
def log(message: str):
    timestamp = datetime.now().strftime("[%Y-%m-%d %H:%M:%S]")
    line = f"{timestamp} {message}"

    print(line)
    os.makedirs(os.path.dirname(LOG_PATH), exist_ok=True)

    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(line + "\n")


# =========================
# CONFIG HANDLING
# =========================
def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_config(cfg):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2)


# =========================
# HELPERS
# =========================
def clean(text: str) -> str:
    return " ".join(str(text or "").lower().split())


def extract_name(description: str) -> str:
    if not description:
        return "No Description"

    desc = clean(description)
    desc = re.sub(r"(overspend|overspent|change)", "", desc)
    desc = re.sub(r"[^a-z\s]", "", desc).strip()

    return desc.title() if desc else "No Description"


def find_row(sheet, name):
    target = clean(name)

    for r in range(2, sheet.max_row + 1):
        if clean(sheet.cell(r, 1).value) == target:
            return r

    return None


def find_column(sheet, month):
    target = clean(month)

    for c in range(1, sheet.max_column + 1):
        val = sheet.cell(1, c).value
        if val and clean(val).startswith(target):
            return c

    return None


def update_cell(sheet, name, month, value):
    row = find_row(sheet, name) or sheet.max_row + 1
    sheet.cell(row, 1).value = name

    col = find_column(sheet, month)
    if not col:
        log(f"Missing column for month '{month}'")
        return None

    prev = sheet.cell(row, col).value or 0
    sheet.cell(row, col).value = value

    return f"{name} [{month}]: {prev} → {value}"


# =========================
# QUICKBOOKS
# =========================
def refresh_token():
    cfg = load_config()

    auth = (
        os.getenv("QB_CLIENT_ID"),
        os.getenv("QB_CLIENT_SECRET")
    )

    payload = {
        "grant_type": "refresh_token",
        "refresh_token": cfg["refresh_token"]
    }

    headers = {"Content-Type": "application/x-www-form-urlencoded"}

    r = requests.post(TOKEN_URL, auth=auth, headers=headers, data=payload)
    r.raise_for_status()

    tokens = r.json()

    if tokens.get("refresh_token"):
        cfg["refresh_token"] = tokens["refresh_token"]
        save_config(cfg)

    return tokens["access_token"]


def fetch_journals(token, realm_id):
    url = f"https://quickbooks.api.intuit.com/v3/company/{realm_id}/query"

    results, start = [], 1
    batch = 1000

    while True:
        query = f"SELECT * FROM JournalEntry STARTPOSITION {start} MAXRESULTS {batch}"

        r = requests.post(
            url,
            headers={"Authorization": f"Bearer {token}"},
            data=query
        )
        r.raise_for_status()

        data = r.json().get("QueryResponse", {}).get("JournalEntry", [])
        if not data:
            break

        results.extend(data)
        start += batch

        if len(data) < batch:
            break

    return results


# =========================
# MAIN PROCESS
# =========================
def main():
    log("Starting automation...")

    token = refresh_token()
    cfg = load_config()

    journals = fetch_journals(token, cfg["realm_id"])
    log(f"Fetched {len(journals)} journal entries")

    wb = load_workbook(EXCEL_PATH)

    month_map = {
        m: m for m in ["Jan","Feb","Mar","Apr","May","Jun",
                       "Jul","Aug","Sep","Oct","Nov","Dec"]
    }

    for sheet_name, account_name in ACCOUNTS:

        if sheet_name not in wb.sheetnames:
            log(f"Missing sheet: {sheet_name}")
            continue

        sheet = wb[sheet_name]
        totals = {}

        for je in journals:
            date = je.get("TxnDate")
            if not date:
                continue

            dt = datetime.strptime(date, "%Y-%m-%d")
            if dt.year != CURRENT_YEAR:
                continue

            month = month_map.get(dt.strftime("%b"))
            if not month:
                continue

            for line in je.get("Line", []):
                acct = clean(line.get("JournalEntryLineDetail", {})
                             .get("AccountRef", {})
                             .get("name", ""))

                if clean(account_name) not in acct:
                    continue

                desc = line.get("Description")
                name = extract_name(desc)
                amount = float(line.get("Amount", 0))

                if "cash in hands" in clean(account_name):
                    d = clean(desc)
                    if "overspend" in d or "overspent" in d:
                        amount = -abs(amount)
                    elif "change" in d:
                        amount = abs(amount)

                key = (name, month)
                totals[key] = totals.get(key, 0) + amount

        updates = [
            update_cell(sheet, name, m, val)
            for (name, m), val in totals.items()
        ]

        updates = [u for u in updates if u]

        log(f"{sheet_name}: {len(updates)} updates")

    wb.save(EXCEL_PATH)
    log("Done.\n")


if __name__ == "__main__":
    main()
